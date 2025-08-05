import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side


class ExcelProcessor:
    def __init__(self, df: pd.DataFrame):
        self.original_df = df

    def filter_data(self):
        df = self.original_df.copy()
        durumlar = df['PAKETLEME DURUMU'].dropna().unique()
        
        kategori_frames = []
        marka_frames = []

        for durum in durumlar:
            df_durum = df[df['PAKETLEME DURUMU'] == durum]

            kategori_df = df_durum.groupby(['MARKA ADI', 'KATEGORİ ADI'], as_index=False).agg(
                SAYFA=('SAYFA', 'sum'),
                KITAP_SAYISI=('KİTAP ID', 'count')
            )
            kategori_df['PAKETLEME DURUMU'] = durum
            kategori_frames.append(kategori_df)

            marka_df = kategori_df.groupby(['MARKA ADI'], as_index=False).agg(
                SAYFA=('SAYFA', 'sum'),
                KITAP_SAYISI=('KITAP_SAYISI', 'sum')
            )
            marka_df['PAKETLEME DURUMU'] = durum
            marka_frames.append(marka_df)

        self.kategori_df = pd.concat(kategori_frames, ignore_index=True)
        self.marka_df = pd.concat(marka_frames, ignore_index=True)

        
        df["KATMAN DURUMU"] = df["KATMANLI PDF BOYUT (MB)"].fillna(0).apply(lambda x: "Katmanlı" if x > 0 else "Katmansız")

        detayli_df = df[
            ['KİTAP ID', 'MARKA ADI', 'KATEGORİ ADI', 'PUBLISHER PREFIX', 'KITAP ADI',
            'SAYFA', 'PAKETLEME DURUMU', 'GELİŞ TARİHİ', 'YÜKLENME TARİHİ',
            'EDİTÖR EKİBİ YÖNETİCİSİ ADI', 'KATMAN DURUMU']
        ].copy()

        self.detayli_df = pd.DataFrame(detayli_df)

        return self.kategori_df, self.marka_df, self.detayli_df
    
    def create_excel(self, sheets: dict[str, pd.DataFrame]) -> bytes:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            wb = writer.book
            thin_border = Border(
                left=Side(style='thin', color='999999'),
                right=Side(style='thin', color='999999'),
                top=Side(style='thin', color='999999'),
                bottom=Side(style='thin', color='999999')
            )
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            row_fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                last_col = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

                for col in ws.columns:
                    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = (max_len + 2) * 1.5

                for row_idx in range(1, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = cell.alignment.copy(horizontal='center')
                        if row_idx == 1:
                            cell.font = Font(bold=True)
                            cell.fill = header_fill
                        else:
                            cell.fill = row_fill_gray if (row_idx - 2) % 2 == 0 else row_fill_white
                            
        

        return output.getvalue()
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        return output.getvalue()


st.set_page_config(page_title="Excel Özetleyici", page_icon="📊")
st.title("📊 Excel Özetleyici ")

uploaded_file = st.file_uploader("📥 Excel dosyasını yükleyin", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    processor = ExcelProcessor(df)
    kategori, marka, detayli= processor.filter_data()

    st.subheader("📌 Kategori")
    st.dataframe(kategori)

    st.subheader("🏷 Marka")
    st.dataframe(marka)

    st.subheader("🔍 Detaylı")
    st.dataframe(detayli)

    excel_bytes = processor.create_excel({
        "Kategori": kategori,
        "Marka": marka,
        "Detaylı": detayli
    })
    st.download_button(
        label="📤 Excel Dosyasını İndir",
        data=excel_bytes,
        file_name="excel_ozet.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
